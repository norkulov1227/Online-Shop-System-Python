from openpyxl import load_workbook, Workbook
from datetime import datetime

class Shop:
    def __init__(self, path, status_file_path) -> None:
        self.path = path
        self.status_file_path = status_file_path
        self.input_field = dict()
        self.data = dict()
        self.report_header = dict()
        self.report_data = dict()

    def read_xlsx(self):
        wb = load_workbook(self.path)
        ws = wb.active

        for i in range(1, ws.max_column+1):
            self.input_field[ws.cell(row = 1, column = i).value] = ws.cell(row = 2, column = i).value
        
        for i in range(1, ws.max_column+1):
            self.data[ws.cell(row = 1, column = i).value] = list()
            for j in range(3, ws.max_row+1):
                self.data.get(ws.cell(row = 1, column = i).value).append(ws.cell(row = j, column = i).value)
    def read_status_xlsx(self):
        wb = load_workbook(self.status_file_path)
        ws = wb.active

        for i in range(1, ws.max_column+1):
            self.report_header[ws.cell(row = 1, column = i).value] = ws.cell(row = 2, column = i).value
        
        for i in range(1, ws.max_column+1):
            self.report_data[ws.cell(row = 1, column = i).value] = list()
            for j in range(3, ws.max_row+1):
                self.report_data.get(ws.cell(row = 1, column = i).value).append(ws.cell(row = j, column = i).value)
            
        




    def get_price(self):
        for i, j in self.input_field.items():
            if i == "narxi":
                return [i, j]
    def get_quantity(self):
        for i, j in self.input_field.items():
            if i == "soni":
                return [i, j]
    def get_name(self):
        for i, j in self.input_field.items():
            if i == "nomi":
                return [i ,j]
    
    
    def add_product(self):  # Mahsulot kiritish
        # print(self.get_data())
        ind = 0
        add_list = dict()
        for key, value in self.input_field.items(): 
            product = input(f"Mahsulot {key} ni {value} turida kiriting: >> ")
            ind += 1
            if product in self.data.get(key):
                if ind == 1:
                    self.sell_add_product(key, product)
                    index = self.data.get(key).index(product)                
                    while True:
                        try:
                            quantity = int(input(f"Kiritilgan mahsulot {self.get_quantity()[0]} ni {self.get_quantity()[1]} turida kirit: >> "))
                            self.sell_add_product('soni', abs(quantity))
                            break
                        except ValueError:
                            print(f"Error: Kiritilgan mahsulot {self.get_quantity()[0]} ni {self.get_quantity()[1]} turida kirit: >> ")
                    while True:
                        try:
                            price = float(input(f"Kiritilgan mahsulot {self.get_price()[0]} ni {self.get_price()[1]} turida kiriti: >>"))
                            self.sell_add_product('narxi', abs(price)*abs(quantity))
                            break
                        except ValueError:
                            print(f"Kiritilgan mahsulot {self.get_price()[0]} ni {self.get_price()[1]} turida kiriti: >>")
                    
                    self.data.get("soni")[index] += abs(quantity)
                    self.data.get("narxi")[index] = abs(price)
                    self.sell_add_product("sanasi", datetime.now().strftime("%Y-%m-%d %X"))
                    self.sell_add_product("status",'add')
                    print(f"\n\nBazaga {product.capitalize()}dan {abs(quantity)} ta qo'shildi\n\n")
                    while True:
                        sorov = input("Bazaga yangi ma'lumot qo'shishni hohlaysizmi? yes/no >> ")                        
                        if sorov == 'yes':
                            self.add_product()
                        elif sorov == 'no':
                            self.main()

            
            if value =='int':
                product = product.lstrip("-")
                product = int(product)
            if key == "narxi":
                product = product.lstrip("-")
                product = float(product)           
                
            self.data.get(key).append(product)
            self.sell_add_product(key, product)
        self.sell_add_product("sanasi", datetime.now().strftime("%Y-%m-%d %X"))
        self.sell_add_product("status",'add')
        self.price_change()
        print("\n\nBazaga yangi mahsulot qo'shildi\n\n")
        while True:
            print("Bazaga yangi ma'lumot qo'shishni hohlaysizmi? yes/no >> ")
            sorov = input(">>>>")
            if sorov == 'yes':
                self.add_product()
            elif sorov == 'no':
                self.main()


    def sell_add_product(self, key, value):  # Report bazaga qoshilgan malumotlarni saqlash
        if key == 'nomi' or key == 'soni' or key == 'narxi' or key == 'sanasi' or key == 'status':
            self.report_data.get(key).append(value)
        
    def price_change(self):
        self.report_data.get("narxi")[-1] *= self.report_data.get("soni")[-1]

    def get_data(self): # Joriy bazadagi mahsulotlarni korish
        if len(self.data.get("nomi")) > 0:
            print("Joriy bazadagi mahsulotlar")
            son = 1
            for i in range(len(self.data.get("nomi"))):
                print(f"{son} - mahsulot")
                for key, value in self.data.items():
                    print(f"{key.capitalize()}: {value[i]}")
                son+=1
                print("\n")
            self.report()
        else: 
            print("Bazada mahsulot yo'q")
    def get_report_add_data(self):  #Bazaga qo'shilgan mahsulotlarni korish
        print("\n       Bazaga qo'shilgan productlar\n")
        if len(self.report_data.get("nomi")) > 0:
            counter = 1
            for i, j in enumerate(self.report_data.get("status")):
                if j == 'add':
                    print(f"{counter} - mahsulot")
                    for key, value in self.report_data.items():
                        print(f"{key.capitalize()}: {value[i]}")
                    print("\n")
                    counter+=1
            self.report()
        else: 
            print("\nMahsulot yo'q\n")
            self.report() 
    
    def get_report_sell_data(self):
        print("\n       Sotilgan productlar\n")
        if len(self.report_data.get("nomi")) > 0:
            counter = 1
            for i, j in enumerate(self.report_data.get("status")):
                if j == 'sell':
                    print(f"{counter} - mahsulot")
                    for key, value in self.report_data.items():
                        print(f"{key.capitalize()}: {value[i]}")
                    print("\n")
                    counter+=1
            self.report()
        else: 
            print("\nMahsulot yo'q\n")
            self.report() 
    
    def sell_product(self): # Mahsulotlarni sotish
        print("\nBazada bor mahsulotlar\nOrqaga qaytish uchun Mahsulot nomiga 0(nol) ni kiriting\n")
        counter = 1
        for key, value in enumerate(self.data.get("nomi")):
            if self.data.get("soni")[key] > 0:
                print(f"{counter}) {value} dan {self.data.get("soni")[key]} ta bor")
                counter += 1
        while True:
            product_name = input("Mahsulot nomini kiriting: >> ")
            if product_name == '0':
                self.main()
            elif product_name not in self.data.get("nomi"):
                print(f"{product_name.capitalize()} nomli mahsulot bazada yo'q.")
            else: break
        while True:
            try:
                quantity = int(input(f"{product_name.capitalize()}dan qancha miqdorda sotib olasiz? >> "))
                ind = self.data.get("nomi").index(product_name)
                if quantity > self.data.get("soni")[ind]:
                    print(f"Bazada {product_name.capitalize()}dan {self.data.get("soni")[ind]} ta bor")
                else:
                    self.data.get("soni")[ind] -= quantity
                    self.sell_add_product("nomi", product_name)
                    self.sell_add_product("soni", quantity)
                    self.sell_add_product("narxi", (self.data.get('narxi')[ind])*quantity + ((self.data.get('narxi')[ind])*quantity)*0.12)
                    self.sell_add_product("sanasi", datetime.now().strftime("%Y-%m-%d %X"))
                    self.sell_add_product("status", 'sell')
                    print("\nMahsulot sotildi!\n")
                    
                    
                    while True:
                        sorov = input("\nYana birorta mahsulot sotib olishni hohlaysizmi? yes/no >> ")                        
                        if sorov == 'yes':
                            self.sell_product()
                        elif sorov == 'no':
                            self.main()
                break
            except ValueError:
                print("\nIltimos mahsulot sonini int turida kiriting!\n")
        # self.delete_empty_products()
        
    def delete_empty_products(self): # Bazada tugagan mahsulotlarni o'chirish
        for i, j in enumerate(self.data.get("soni")):
            if j == 0:
                for key, value in self.data.items():
                    self.data.get(key).remove(value[i])


    def report(self): #hisobot bolimidan bazalarni tanlash
        self.delete_empty_products()
        sorov = input("\nJoriy bazani korish >> 1\nBazaga qo'shilgan productlarni ko'rish >> 2\nSotilgan productlarni ko'rish >> 3\nQaytish uchun hohlagan buyruq >> ")
        if sorov == '1':
            self.get_data()
        elif sorov == '2':
            self.get_report_add_data()
        elif sorov == '3':
            self.get_report_sell_data()
        else: self.main()
    def exit_and_save(self):    # Yakun! Bu ma'lumotlarni faylga saqlaydi va dastur ishdan toxtaydi
        wb = Workbook()
        ws = wb.active

        ws.append(list(self.input_field.keys()))
        ws.append(list(self.input_field.values()))
        for i in range(len(self.data.get("nomi"))):
            save_list = list()
            for key, value in self.data.items():
                save_list.append(value[i])
            ws.append(save_list)
        wb.save("data.xlsx")

        wb1 = Workbook()
        ws1 = wb1.active

        ws1.append(list(self.report_header.keys()))
        ws1.append(list(self.report_header.values()))

        for i in range(len(self.report_data.get("nomi"))):
            save_report_list = list()
            for key, value in self.report_data.items():
                save_report_list.append(value[i])
            ws1.append(save_report_list)
        wb1.save("status.xlsx")
        quit()

    def main(self):  #Main asosiy qism
        while True:
            print("Mahsulot qo'shish >> 1\nMahsulot sotish >> 2\nHisobot >> 3\nEXIT AND SAVE >> 0")
            sorov = input("Bo'limni tanlang: >> ")
            if sorov == '0':
                print("Siz tizimdan chiqdingiz.")
                self.exit_and_save()
            elif sorov == '1':
                self.add_product()
            elif sorov == '2':
                self.sell_product()
            elif sorov == '3':
                self.report()
        